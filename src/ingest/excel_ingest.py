"""
Excel Ingest Module for Theophysics Definition System

Uses established packages:
- openpyxl: For .xlsx file reading (44M+ weekly downloads)
- pandas: For data manipulation (300M+ weekly downloads)
- xlrd: For legacy .xls files (15M+ weekly downloads)

Features:
- Multi-sheet processing
- Header detection
- Type inference
- Source attribution (marks all data as EXCEL source)
- Batch processing for large files
- Progress tracking with tqdm
"""

import os
import hashlib
from pathlib import Path
from typing import Optional, List, Dict, Any, Generator, Tuple
from datetime import datetime
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# Local imports
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from db.schema import (
    SourceType, ConfidenceLevel, IngestSession, IngestRecord,
    ExcelSheet, get_session, get_engine
)


@dataclass
class ExcelCell:
    """Represents a single Excel cell with metadata"""
    value: Any
    row: int
    column: int
    cell_ref: str  # e.g., "A1"
    data_type: str
    is_formula: bool = False
    formula: Optional[str] = None


@dataclass
class ExcelRow:
    """Represents a row of Excel data"""
    row_number: int
    cells: Dict[str, Any]  # column_name -> value
    raw_cells: List[ExcelCell] = field(default_factory=list)


@dataclass
class ExcelSheetData:
    """Represents data from an Excel sheet"""
    file_path: str
    sheet_name: str
    headers: List[str]
    rows: List[ExcelRow]
    row_count: int
    column_count: int
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelIngester:
    """
    Ingests Excel files into PostgreSQL with full source attribution.

    All ingested data is marked with:
    - source_type: EXCEL
    - source_file: path to the Excel file
    - source_sheet: sheet name
    - source_row: row number
    - source_cell: cell reference

    Usage:
        ingester = ExcelIngester(db_connection_string)
        ingester.ingest_file("data.xlsx")

        # Or for batch processing:
        ingester.ingest_directory("/path/to/excel/files")
    """

    def __init__(self, db_connection_string: Optional[str] = None, batch_size: int = 100):
        """
        Initialize the Excel ingester.

        Args:
            db_connection_string: PostgreSQL connection string
            batch_size: Number of rows to process before committing
        """
        self.db_connection_string = db_connection_string
        self.batch_size = batch_size
        self.engine = None
        self.session = None
        self.current_ingest_session: Optional[IngestSession] = None

        if db_connection_string:
            self.engine = get_engine(db_connection_string)

    def _ensure_session(self):
        """Ensure database session exists"""
        if self.engine and not self.session:
            self.session = get_session(self.engine)

    def _create_ingest_session(self, source_path: str) -> IngestSession:
        """Create a new ingest session for tracking"""
        self._ensure_session()
        session = IngestSession(
            source_type=SourceType.EXCEL,
            source_path=source_path,
            source_name=Path(source_path).name,
            metadata={"ingester": "ExcelIngester", "version": "1.0"}
        )
        if self.session:
            self.session.add(session)
            self.session.commit()
        return session

    def _hash_content(self, content: str) -> str:
        """Create SHA-256 hash of content for deduplication"""
        return hashlib.sha256(content.encode('utf-8')).hexdigest()

    def read_file(self, file_path: str) -> List[ExcelSheetData]:
        """
        Read an Excel file and return structured data from all sheets.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of ExcelSheetData objects, one per sheet
        """
        file_path = str(Path(file_path).resolve())
        sheets_data = []

        # Determine file type and use appropriate reader
        if file_path.endswith('.xlsx'):
            sheets_data = self._read_xlsx(file_path)
        elif file_path.endswith('.xls'):
            sheets_data = self._read_xls(file_path)
        else:
            # Try pandas for other formats
            sheets_data = self._read_with_pandas(file_path)

        return sheets_data

    def _read_xlsx(self, file_path: str) -> List[ExcelSheetData]:
        """Read .xlsx file using openpyxl for full fidelity"""
        sheets_data = []
        workbook = load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Get dimensions
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0

            if max_row == 0 or max_col == 0:
                continue

            # Extract headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col)
                header = str(cell.value) if cell.value else f"Column_{col}"
                headers.append(header)

            # Extract rows
            rows = []
            for row_num in range(2, max_row + 1):
                cells_dict = {}
                raw_cells = []

                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell_ref = f"{get_column_letter(col_num)}{row_num}"

                    excel_cell = ExcelCell(
                        value=cell.value,
                        row=row_num,
                        column=col_num,
                        cell_ref=cell_ref,
                        data_type=cell.data_type if hasattr(cell, 'data_type') else 'unknown',
                        is_formula=str(cell.value).startswith('=') if cell.value else False
                    )

                    raw_cells.append(excel_cell)
                    cells_dict[header] = cell.value

                rows.append(ExcelRow(
                    row_number=row_num,
                    cells=cells_dict,
                    raw_cells=raw_cells
                ))

            sheets_data.append(ExcelSheetData(
                file_path=file_path,
                sheet_name=sheet_name,
                headers=headers,
                rows=rows,
                row_count=max_row - 1,  # Exclude header
                column_count=max_col,
                metadata={
                    "engine": "openpyxl",
                    "file_format": "xlsx"
                }
            ))

        workbook.close()
        return sheets_data

    def _read_xls(self, file_path: str) -> List[ExcelSheetData]:
        """Read .xls file using pandas with xlrd backend"""
        return self._read_with_pandas(file_path, engine='xlrd')

    def _read_with_pandas(self, file_path: str, engine: Optional[str] = None) -> List[ExcelSheetData]:
        """Read Excel file using pandas (works for .xls and .xlsx)"""
        sheets_data = []

        # Read all sheets
        xl = pd.ExcelFile(file_path, engine=engine)

        for sheet_name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet_name)

            if df.empty:
                continue

            headers = df.columns.tolist()

            rows = []
            for idx, row in df.iterrows():
                row_num = idx + 2  # Account for header and 0-indexing
                cells_dict = row.to_dict()

                raw_cells = []
                for col_num, (header, value) in enumerate(cells_dict.items(), 1):
                    cell_ref = f"{get_column_letter(col_num)}{row_num}"
                    raw_cells.append(ExcelCell(
                        value=value if pd.notna(value) else None,
                        row=row_num,
                        column=col_num,
                        cell_ref=cell_ref,
                        data_type=type(value).__name__ if pd.notna(value) else 'null'
                    ))

                # Clean NaN values
                cells_dict = {k: (v if pd.notna(v) else None) for k, v in cells_dict.items()}

                rows.append(ExcelRow(
                    row_number=row_num,
                    cells=cells_dict,
                    raw_cells=raw_cells
                ))

            sheets_data.append(ExcelSheetData(
                file_path=file_path,
                sheet_name=sheet_name,
                headers=headers,
                rows=rows,
                row_count=len(df),
                column_count=len(headers),
                metadata={
                    "engine": "pandas",
                    "file_format": Path(file_path).suffix.lstrip('.')
                }
            ))

        return sheets_data

    def ingest_file(self, file_path: str, progress: bool = True) -> Dict[str, Any]:
        """
        Ingest an Excel file into the database.

        Args:
            file_path: Path to the Excel file
            progress: Show progress bar

        Returns:
            Dictionary with ingest statistics
        """
        file_path = str(Path(file_path).resolve())

        # Create ingest session
        self.current_ingest_session = self._create_ingest_session(file_path)

        stats = {
            "file": file_path,
            "sheets_processed": 0,
            "rows_processed": 0,
            "records_created": 0,
            "errors": []
        }

        try:
            sheets_data = self.read_file(file_path)

            for sheet_data in sheets_data:
                sheet_stats = self._ingest_sheet(sheet_data, progress)
                stats["sheets_processed"] += 1
                stats["rows_processed"] += sheet_stats["rows"]
                stats["records_created"] += sheet_stats["created"]
                stats["errors"].extend(sheet_stats["errors"])

        except Exception as e:
            stats["errors"].append(f"File error: {str(e)}")

        # Update ingest session
        if self.current_ingest_session and self.session:
            self.current_ingest_session.completed_at = datetime.utcnow()
            self.current_ingest_session.records_processed = stats["rows_processed"]
            self.current_ingest_session.records_created = stats["records_created"]
            self.current_ingest_session.records_failed = len(stats["errors"])
            self.current_ingest_session.error_log = "\n".join(stats["errors"]) if stats["errors"] else None
            self.session.commit()

        return stats

    def _ingest_sheet(self, sheet_data: ExcelSheetData, progress: bool) -> Dict[str, Any]:
        """Ingest a single Excel sheet"""
        stats = {"rows": 0, "created": 0, "errors": []}

        self._ensure_session()

        # Record the sheet
        if self.session:
            excel_sheet = ExcelSheet(
                file_path=sheet_data.file_path,
                file_name=Path(sheet_data.file_path).name,
                sheet_name=sheet_data.sheet_name,
                row_count=sheet_data.row_count,
                column_count=sheet_data.column_count,
                headers=sheet_data.headers,
                session_id=self.current_ingest_session.id if self.current_ingest_session else None
            )
            self.session.add(excel_sheet)

        # Process rows
        rows_iter = tqdm(sheet_data.rows, desc=f"Sheet: {sheet_data.sheet_name}") if progress else sheet_data.rows

        batch = []
        for row in rows_iter:
            try:
                # Create ingest record for each row
                content = str(row.cells)
                record = IngestRecord(
                    session_id=self.current_ingest_session.id if self.current_ingest_session else None,
                    source_type=SourceType.EXCEL,
                    source_file=sheet_data.file_path,
                    source_sheet=sheet_data.sheet_name,
                    source_row=row.row_number,
                    raw_content=content,
                    processed_content=content,
                    content_hash=self._hash_content(content),
                    confidence=ConfidenceLevel.HIGH,  # Excel data is reliable
                    needs_review=False  # Mark as not needing review since it's from structured source
                )

                batch.append(record)
                stats["rows"] += 1
                stats["created"] += 1

                # Batch commit
                if len(batch) >= self.batch_size and self.session:
                    self.session.add_all(batch)
                    self.session.commit()
                    batch = []

            except Exception as e:
                stats["errors"].append(f"Row {row.row_number}: {str(e)}")

        # Commit remaining batch
        if batch and self.session:
            self.session.add_all(batch)
            self.session.commit()

        if self.session:
            excel_sheet.fully_processed = True
            self.session.commit()

        return stats

    def ingest_directory(self, directory_path: str, recursive: bool = True,
                         extensions: List[str] = None) -> Dict[str, Any]:
        """
        Ingest all Excel files from a directory.

        Args:
            directory_path: Path to directory
            recursive: Search subdirectories
            extensions: File extensions to process (default: ['.xlsx', '.xls'])

        Returns:
            Aggregated statistics
        """
        if extensions is None:
            extensions = ['.xlsx', '.xls']

        directory = Path(directory_path)
        pattern = "**/*" if recursive else "*"

        files = []
        for ext in extensions:
            files.extend(directory.glob(f"{pattern}{ext}"))

        total_stats = {
            "files_processed": 0,
            "total_sheets": 0,
            "total_rows": 0,
            "total_records": 0,
            "errors": []
        }

        for file_path in tqdm(files, desc="Processing Excel files"):
            try:
                stats = self.ingest_file(str(file_path), progress=False)
                total_stats["files_processed"] += 1
                total_stats["total_sheets"] += stats["sheets_processed"]
                total_stats["total_rows"] += stats["rows_processed"]
                total_stats["total_records"] += stats["records_created"]
                total_stats["errors"].extend(stats["errors"])
            except Exception as e:
                total_stats["errors"].append(f"{file_path}: {str(e)}")

        return total_stats

    def to_dataframe(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Read Excel file directly to pandas DataFrame (convenience method).

        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to read (default: first sheet)

        Returns:
            pandas DataFrame
        """
        if sheet_name:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        return pd.read_excel(file_path)

    def get_sheet_preview(self, file_path: str, sheet_name: str = None,
                          n_rows: int = 10) -> Dict[str, Any]:
        """
        Get a preview of an Excel sheet.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet to preview (default: first sheet)
            n_rows: Number of rows to preview

        Returns:
            Dictionary with headers and sample rows
        """
        sheets = self.read_file(file_path)

        if not sheets:
            return {"error": "No sheets found"}

        sheet = sheets[0] if not sheet_name else next(
            (s for s in sheets if s.sheet_name == sheet_name), sheets[0]
        )

        return {
            "file": file_path,
            "sheet": sheet.sheet_name,
            "headers": sheet.headers,
            "total_rows": sheet.row_count,
            "preview_rows": [
                {
                    "row_number": row.row_number,
                    "data": row.cells
                }
                for row in sheet.rows[:n_rows]
            ],
            "source_attribution": {
                "source_type": "EXCEL",
                "source_file": file_path,
                "source_sheet": sheet.sheet_name,
                "ingested_by": "python/ExcelIngester"
            }
        }

    def close(self):
        """Close database session"""
        if self.session:
            self.session.close()
            self.session = None


# === CONVENIENCE FUNCTIONS ===

def quick_ingest(file_path: str, db_url: str = None) -> Dict[str, Any]:
    """
    Quick one-liner to ingest an Excel file.

    Args:
        file_path: Path to Excel file
        db_url: PostgreSQL connection string (optional)

    Returns:
        Ingest statistics
    """
    ingester = ExcelIngester(db_url)
    result = ingester.ingest_file(file_path)
    ingester.close()
    return result


def excel_to_dict(file_path: str) -> Dict[str, List[Dict]]:
    """
    Convert Excel file to dictionary (no database required).

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with sheet names as keys, list of row dicts as values
    """
    ingester = ExcelIngester()
    sheets = ingester.read_file(file_path)

    result = {}
    for sheet in sheets:
        result[sheet.sheet_name] = [
            {
                **row.cells,
                "__source__": {
                    "type": "EXCEL",
                    "file": file_path,
                    "sheet": sheet.sheet_name,
                    "row": row.row_number
                }
            }
            for row in sheet.rows
        ]

    return result
